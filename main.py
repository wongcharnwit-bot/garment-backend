from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import io
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = FastAPI(title="Garment Auto Balancer – Dynamic Core")

app.add_middleware(
    CORSMiddleware, 
    allow_origins=["*"], 
    allow_methods=["*"], 
    allow_headers=["*"]
)

# --- HEALTH CHECK (Wake Up & Status) ---
@app.get("/")
def health_check():
    return {"status": "ok", "message": "Garment Balancer Backend is Ready 🚀"}

# --- CORE ALGORITHM (Your Logic) ---
def true_thai_balancing(processes, num_operators, time_key="smv"):
    """ 
    Line Balancing (Sequential Flow / Water Flow)
    Updated: Uses 'time_key' to decide if we balance based on SMV or CT
    """
    if num_operators <= 0:
        return []

    # Calculate Target Cycle Time (Takt Time) based on SELECTED time
    total_time = sum(p[time_key] for p in processes)
    target_ct = total_time / num_operators if num_operators > 0 else 0
    
    operators = [{"op": i+1, "sec": 0.0, "tasks": []} for i in range(num_operators)] # Store op as INT index first
    
    current_op_idx = 0
    
    for proc in processes:
        remaining_proc_time = proc[time_key] # Use dynamic key
        proc_no = proc["no"]
        proc_desc = proc["desc"]
        original_time = proc[time_key] # Use dynamic key

        # Keep distributing this process until it's finished
        while remaining_proc_time > 0.001:
            
            # Safety: If we run out of operators, dump everything to the last one
            if current_op_idx >= num_operators:
                current_op_idx = num_operators - 1
            
            current_op = operators[current_op_idx]
            
            # How much space is left for this operator?
            space_left = target_ct - current_op["sec"]
            
            # If this operator is already full (or overfilled slightly), move to next
            if space_left <= 0.001:
                current_op_idx += 1
                continue
            
            # Determine how much this operator can take
            take_time = min(remaining_proc_time, space_left)
            
            # Calculate Percentage for display
            percentage = (take_time / original_time) * 100 if original_time > 0 else 0
            
            # Format Description
            part_str = f' - "{proc["part"]}"' if proc.get("part") else ""
            task_desc = f"No.{proc_no}: {proc_desc}"
            if percentage < 99.9:
                task_desc += f" ({percentage:.0f}%)"
            
            # Append Part at the very end
            task_desc += part_str
            
            # Add task to operator
            current_op["tasks"].append({
                "no": proc_no,
                "desc": task_desc,
                "time": take_time,
                "percentage": percentage
            })
            
            # Update counters
            current_op["sec"] += take_time
            remaining_proc_time -= take_time
            
            # If we just filled this operator perfectly (or close to it), 
            # and there is still work left in this process, move to next operator
            if remaining_proc_time > 0.001:
                current_op_idx += 1

    return operators

async def parse_excel_structure(file_content):
    """
    Reads Excel and returns an ORDERED dictionary.
    Updated: Reads Col G (Index 6) as SMV, Col H (Index 7) as CT, Col I (Index 8) as Part.
    """
    try:
        df = pd.read_excel(io.BytesIO(file_content), sheet_name="PA sheet", header=None, engine='openpyxl')
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error reading Excel: {str(e)}")

    sections = {} # Use dict to keep insertion order
    
    # Start from Row 5 (Index 4)
    for i in range(4, df.shape[0]):
        row = df.iloc[i]
        try:
            # Column G (Index 6) is SMV Time
            smv_val = pd.to_numeric(row.iloc[6], errors='coerce')
            if pd.isna(smv_val) or smv_val <= 0: continue
            
            # NEW: Column H (Index 7) is CT Time (formerly NGIE)
            # If CT is blank or 0, we fallback to SMV value
            ct_val = pd.to_numeric(row.iloc[7], errors='coerce')
            if pd.isna(ct_val) or ct_val <= 0:
                ct_val = 0 # Explicitly 0 if missing, don't fallback to SMV unless logic requires (User said summary box should be zero if missing)
            
            # NEW: Column I (Index 8) is Part Name
            part_name = str(row.iloc[8]).strip() if pd.notna(row.iloc[8]) else ""

            # Column C (Index 2) is Section Name
            raw_sec_name = str(row.iloc[2]).strip()
            if not raw_sec_name or raw_sec_name.lower() == 'nan': continue
            
            # Store both time values + Flow/MC for Export
            proc = {
                "no": int(row.iloc[1]) if pd.notna(row.iloc[1]) else i,
                "desc": str(row.iloc[5]).strip(),
                "flow": str(row.iloc[3]) if pd.notna(row.iloc[3]) else "", # Captured for Export
                "mc": str(row.iloc[4]) if pd.notna(row.iloc[4]) else "",   # Captured for Export
                "smv": float(smv_val),
                "ct": float(ct_val),
                "part": part_name
            }
            
            if raw_sec_name not in sections:
                sections[raw_sec_name] = []
            
            sections[raw_sec_name].append(proc)
            
        except Exception:
            continue
            
    return sections

# --- NEW: EXCEL GENERATION LOGIC ---
def generate_excel_report(sections_data, balanced_data_list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Balancing Report"
    
    # Styles
    header_fill = PatternFill(start_color="36454F", end_color="36454F", fill_type="solid") # Dark Gray
    section_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") # Light Gray
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 1. Setup Static Headers (Row 5)
    static_headers = ["No", "PPA", "Flow", "MC", "Process", "SMV", "CT", "Part"]
    for col_num, header in enumerate(static_headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # 2. Setup Dynamic Operator Headers (Row 4 & 5)
    current_col = 9 # Starts after Part (Col I is 8)
    
    op_task_map = {} # Map: task_no -> { global_col_index: time }

    for sec_data in balanced_data_list:
        sec_name = sec_data['name']
        ops = sec_data['operators']
        num_ops = len(ops)
        
        # Merge Header for Section (Row 4)
        if num_ops > 0:
            start_col = current_col
            end_col = current_col + num_ops - 1
            ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=end_col)
            sec_cell = ws.cell(row=4, column=start_col, value=sec_name)
            sec_cell.alignment = center_align
            sec_cell.font = Font(bold=True)
            sec_cell.fill = section_fill
            
            # Operator Headers (Row 5)
            # Reset numbering for each section (Op 1, Op 2...)
            for i, op in enumerate(ops):
                col_idx = current_col + i
                cell = ws.cell(row=5, column=col_idx, value=f"Op {i+1}")
                cell.alignment = center_align
                cell.border = thin_border
                cell.font = Font(bold=True)
                
                # Map tasks for this operator to the global column index
                for task in op['tasks']:
                    t_no = task['no']
                    if t_no not in op_task_map: op_task_map[t_no] = {}
                    op_task_map[t_no][col_idx] = task['time']
            
            current_col += num_ops

    # 3. Fill Data Rows (Using original sequential order)
    current_row = 6
    for sec_name, procs in sections_data.items():
        for proc in procs:
            # Static Data
            ws.cell(row=current_row, column=1, value=proc['no']).border = thin_border
            ws.cell(row=current_row, column=2, value=sec_name).border = thin_border
            ws.cell(row=current_row, column=3, value=proc['flow']).border = thin_border
            ws.cell(row=current_row, column=4, value=proc['mc']).border = thin_border
            ws.cell(row=current_row, column=5, value=proc['desc']).border = thin_border
            ws.cell(row=current_row, column=6, value=proc['smv']).border = thin_border
            ws.cell(row=current_row, column=7, value=proc['ct']).border = thin_border
            ws.cell(row=current_row, column=8, value=proc['part']).border = thin_border
            
            # Dynamic Operator Data
            assignments = op_task_map.get(proc['no'], {})
            
            # Iterate through all operator columns created
            total_cols = current_col - 1
            for col_idx in range(9, total_cols + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = thin_border
                if col_idx in assignments:
                    cell.value = assignments[col_idx]
                    cell.alignment = center_align
                    # Highlight split tasks (red text)
                    if len(assignments) > 1:
                        cell.font = Font(color="FF0000") 
            
            current_row += 1

    # Adjust Column Widths
    ws.column_dimensions['E'].width = 40 # Process Desc
    for col in range(8, current_col):
        ws.column_dimensions[get_column_letter(col)].width = 8

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream

# --- ENDPOINTS ---

@app.post("/analyze")
async def analyze_file(
    total_operators: int = Form(...),
    file: UploadFile = File(...)
):
    """
    # Step 1: Calculate Takt Time & Suggestions for BOTH SMV and CT modes.
    """
    content = await file.read()
    sections_data = await parse_excel_structure(content)
    
    if not sections_data:
        raise HTTPException(status_code=400, detail="No valid data found in Excel")

    # Helper function to calculate stats for a specific time mode (smv or ct)
    def calculate_stats(mode_key):
        total_mode = sum(sum(p[mode_key] for p in procs) for procs in sections_data.values())
        takt_time_mode = total_mode / total_operators if total_operators > 0 else 0
        
        results = []
        for sec_name, procs in sections_data.items():
            sec_total = sum(p[mode_key] for p in procs)
            theoretical = sec_total / takt_time_mode if takt_time_mode > 0 else 0
            suggested = max(1, round(theoretical))
            
            results.append({
                "name": sec_name,
                "total": round(sec_total, 2), # Total time for this section
                "theoretical": round(theoretical, 2), 
                "suggested": suggested 
            })
        return {"total_time": round(total_mode, 2), "takt_time": round(takt_time_mode, 2), "sections": results}

    # Return both datasets so frontend can toggle
    return {
        "smv_data": calculate_stats("smv"),
        "ct_data": calculate_stats("ct")
    }

@app.post("/balance")
async def balance_dynamic(
    config_str: str = Form(...), 
    time_mode: str = Form(...), # "smv" or "ct"
    selected_sections_str: str = Form("[]"), # JSON list of selected sections
    file: UploadFile = File(...)
):
    """
    Step 2: Balance based on time_mode, but calculate Efficiency for BOTH.
    """
    try:
        config = json.loads(config_str) 
    except:
        raise HTTPException(status_code=400, detail="Invalid configuration format")

    content = await file.read()
    sections_data = await parse_excel_structure(content)
    
    results_list = []
    all_ops_for_stats = []
    
    for sec_name, procs in sections_data.items():
        num_ops = int(config.get(sec_name, 1))
        
        # 1. Balance using the SELECTED Mode
        # Returns simple list: [{"op": 1, "sec": 10.5, "tasks":...}]
        balanced_ops_raw = true_thai_balancing(procs, num_ops, time_key=time_mode)
        
        # Format "op" key to "Op 1", "Op 2" for display
        balanced_ops = []
        for op in balanced_ops_raw:
             op_display = op.copy()
             op_display["op"] = f"Op {op['op']}"
             balanced_ops.append(op_display)

        # 2. Calculate Totals for BOTH modes (for efficiency calc)
        sec_total_smv = sum(p["smv"] for p in procs)
        sec_total_ct = sum(p["ct"] for p in procs)
        
        # 3. Find Bottleneck (based on the balanced result)
        sec_bn = max((op["sec"] for op in balanced_ops), default=0)

        # 4. Calculate Metrics
        sec_output = round(3600 / sec_bn, 0) if sec_bn > 0 else 0
        
        # Efficiency Formula: (Total Time / (Bottleneck * Ops)) * 100
        denom = sec_bn * num_ops
        sec_eff_smv = round((sec_total_smv * 100) / denom, 1) if denom > 0 else 0
        sec_eff_ct = round((sec_total_ct * 100) / denom, 1) if denom > 0 else 0
        
        # 5. Color Logic
        for op in balanced_ops:
            if op["sec"] == sec_bn and sec_bn > 0:
                op["color"] = "orange"
            else:
                op["color"] = "green"
        
        results_list.append({
            "name": sec_name,
            "total_time_used": round(sum(p[time_mode] for p in procs), 2), 
            "operators": balanced_ops,
            "section_bn": sec_bn,
            "sec_output": sec_output,
            "sec_eff_smv": sec_eff_smv,   
            "sec_eff_ct": sec_eff_ct,
            "num_ops": num_ops # Store for Global Calc
        })
        
        all_ops_for_stats.extend(balanced_ops)

    # --- Global Analytics ---
    all_times = [op["sec"] for op in all_ops_for_stats if op["sec"] > 0]
    line_bottleneck = max(all_times) if all_times else 0
    
    # Update Red Color for Global Bottleneck
    for section in results_list:
        for op in section["operators"]:
            if op["sec"] == line_bottleneck:
                op["color"] = "red"
    
    total_man_global = sum(len(sec["operators"]) for sec in results_list)
    global_denom = line_bottleneck * total_man_global
    
    # Calculate Global Totals for both
    global_total_smv = sum(sum(p["smv"] for p in procs) for procs in sections_data.values())
    
    output = round(3600 / line_bottleneck, 0) if line_bottleneck > 0 else 0
    
    # Global Efficiencies
    eff_smv_global = round((global_total_smv * 100) / global_denom, 1) if global_denom > 0 else 0
    
    # --- NEW: % LINE BALANCE CALCULATION (Weighted Average Cycle Time) ---
    # Formula: (Sum of Section Cycle Times) / (Max Seq Cycle Time * Count) * 100
    
    selected_sections = []
    try:
        if 'selected_sections_str' in locals():
             selected_sections = json.loads(selected_sections_str)
    except:
        selected_sections = []

    if not selected_sections:
        selected_sections = list(sections_data.keys())

    # Get Cycle Times for Selected Sections
    # Cycle Time = Total Time Used / Num Ops
    
    cycle_times = []
    for sec_name in selected_sections:
        # Find the result for this section
        sec_res = next((r for r in results_list if r["name"] == sec_name), None)
        if sec_res:
            # Re-calculate cycle time based on the Balanced output
            # Actually, "Cycle Time" per person is what we balanced. 
            # But the formula requested: "P2 total time use = 383 sec ==> After balance = 76.60 sec/person"
            # So Cycle Time = 76.60.
            
            # Use the "section_bn" (Bottleneck) as the Cycle Time for that section?
            # User example: "P2 total time 383 => After balance 76.60". 383/5 = 76.6. 
            # So Cycle Time = Total / Ops. 
            # NOTE: sec_res["section_bn"] is the ACTUAL max time an operator takes. 
            # Theoretical Cycle Time = Total / Ops. 
            # User example 2: "Ass total time 263.81 => After balance 87.94". 263.81 / 3 = 87.936.
            # So "After balance" = Total / Ops (Theoretical Average? or Max of balanced?)
            
            # Wait, 76.60 * 5 = 383. 87.94 * 3 = 263.82.
            # Using Total / Ops is consistent with the example.
            
            t_used = sec_res["total_time_used"]
            n_ops = sec_res["num_ops"]
            
            cycle_t = t_used / n_ops if n_ops > 0 else 0
            cycle_times.append(cycle_t)
            
    # Calculate Line Balance %
    if cycle_times:
        sum_cycles = sum(cycle_times)
        max_cycle = max(cycle_times)
        count_cycles = len(cycle_times)
        denom_lb = max_cycle * count_cycles
        
        line_balance_eff = round((sum_cycles * 100) / denom_lb, 2) if denom_lb > 0 else 0
    else:
        line_balance_eff = 0

    # --- AI Suggestion ---
    suggestions = []
    
    if line_balance_eff < 60:
        suggestions.append(f"⚠️ Line Balance is low ({line_balance_eff}%). Consider combining processes.")
    elif line_balance_eff > 90:
        suggestions.append(f"✅ Excellent Line Balance ({line_balance_eff}%).")
    
    if line_bottleneck > 0:
        bn_secs = [res["name"] for res in results_list if res["section_bn"] == line_bottleneck]
        # Clean naming for suggestion
        formatted_bn = [f"{s}" for s in bn_secs] 
        loc_str = " & ".join(formatted_bn)
        suggestions.append(f"🛑 Critical Bottleneck: {line_bottleneck}s in {loc_str}.")

        if any("ass" in s.lower() for s in bn_secs):
             suggestions.append(f"👉 Action: Check Assembly machines.")

    return {
        "bottleneck": round(line_bottleneck, 2),
        "output": output,
        "eff_smv": eff_smv_global,   
        "line_balance_eff": line_balance_eff,
        "suggest": " ".join(suggestions),
        "sections_results": results_list
    }

# --- NEW: EXPORT ENDPOINT ---
@app.post("/export")
async def export_excel(
    config_str: str = Form(...), 
    time_mode: str = Form(...), 
    file: UploadFile = File(...)
):
    try:
        config = json.loads(config_str) 
    except:
        raise HTTPException(status_code=400, detail="Invalid configuration format")
    
    content = await file.read()
    sections_data = await parse_excel_structure(content)
    
    # 1. Run Balancing Logic (Internal - same as /balance)
    balanced_results = []
    for sec_name, procs in sections_data.items():
        num_ops = int(config.get(sec_name, 1))
        balanced_ops = true_thai_balancing(procs, num_ops, time_key=time_mode)
        balanced_results.append({'name': sec_name, 'operators': balanced_ops})
    
    # 2. Generate Excel
    excel_file = generate_excel_report(sections_data, balanced_results)
    
    return StreamingResponse(
        excel_file, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename=Balancing_Report_{time_mode.upper()}.xlsx"}
    )